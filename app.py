from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
import json

app = Flask(__name__)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'})

    if file:
        wb = load_workbook(file.stream)
        ws = wb.active
        data = []
        bank_name = request.form.get('bankName', '')

        for row in ws.iter_rows(min_row=2, values_only=True):  # Start from the second row to skip headers
            date, voucher_type, voucher_number, dr_cr, party_details, debit_ledger, credit_ledger, amount_single, debit_amount, credit_amount, narration = row

            if not voucher_type:
                if dr_cr and dr_cr.strip().lower() in ['dr', 'dr.', 'debit']:
                    voucher_type = 'Payment'
                elif dr_cr and dr_cr.strip().lower() in ['cr', 'cr.', 'credit']:
                    voucher_type = 'Receipt'
                elif not dr_cr:
                    if debit_amount:
                        voucher_type = 'Payment'
                    elif not debit_amount:
                        voucher_type = 'Receipt'

            if not dr_cr:
                if debit_amount:
                    dr_cr = 'DR'
                elif credit_amount:
                    dr_cr = 'CR'

            if not party_details:
                party_details = 'XXX'

            if voucher_type == 'Payment':
                debit_ledger = party_details
                credit_ledger = bank_name
            elif voucher_type == 'Receipt':
                debit_ledger = bank_name
                credit_ledger = party_details

            if not amount_single:
                if debit_amount:
                    amount_single = debit_amount
                elif credit_amount:
                    amount_single = credit_amount

            if not debit_amount and not credit_amount:
                if voucher_type == 'Payment':
                    debit_amount = amount_single
                elif voucher_type == 'Receipt':
                    credit_amount = amount_single

            data.append({
                "Date": date if date else "",
                "Voucher Type": voucher_type if voucher_type else "",
                "Voucher Number": voucher_number if voucher_number else "",
                "DR/CR": dr_cr if dr_cr else "",
                "Party Details": party_details if party_details else "",
                "Debit Ledger": debit_ledger if debit_ledger else "",
                "Credit Ledger": credit_ledger if credit_ledger else "",
                "Amount in Single Column": amount_single if amount_single else "",
                "Debit Amount": debit_amount if debit_amount else "",
                "Credit Amount": credit_amount if credit_amount else "",
                "Narration": narration if narration else ""
            })

        return jsonify(data)


@app.route('/generate_xml', methods=['POST'])
def generate_xml_route():
    try:
        data = json.loads(request.data)
        xml_data = "<ENVELOPE>\n<HEADER>\n<TALLYREQUEST>Import Data</TALLYREQUEST>\n</HEADER>\n<BODY>\n<IMPORTDATA>\n<REQUESTDESC>\n<REPORTNAME>Vouchers</REPORTNAME>\n</REQUESTDESC>\n<REQUESTDATA>\n"
        for entry in data:
            try:
                amount = float(entry['Amount in Single Column'])
                xml_data += f"<TALLYMESSAGE><VOUCHER VCHTYPE=\"{entry['Voucher Type']}\" ACTION=\"Create\"><DATE>{entry['Date']}</DATE><VOUCHERTYPENAME>{entry['Voucher Type']}</VOUCHERTYPENAME>"
                if entry['Voucher Number']:
                    xml_data += f"<VOUCHERNUMBER>{entry['Voucher Number']}</VOUCHERNUMBER>"
                if entry['Narration']:
                    xml_data += f"<NARRATION>{entry['Narration']}</NARRATION>"
                xml_data += f"<ALLLEDGERENTRIES.LIST><LEDGERNAME>{entry['Debit Ledger']}</LEDGERNAME><ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE><AMOUNT>{-amount}</AMOUNT></ALLLEDGERENTRIES.LIST><ALLLEDGERENTRIES.LIST><LEDGERNAME>{entry['Credit Ledger']}</LEDGERNAME><ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE><AMOUNT>{amount}</AMOUNT></ALLLEDGERENTRIES.LIST></VOUCHER></TALLYMESSAGE>\n"
            except ValueError:
                return jsonify({'error': 'Amount value could not be converted to float'})
        xml_data += "</REQUESTDATA>\n</IMPORTDATA>\n</BODY>\n</ENVELOPE>\n"
        return xml_data
    except Exception as e:
        return jsonify({'error': str(e)})


if __name__ == '__main__':
    app.run(debug=True)
